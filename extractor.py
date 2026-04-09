"""
Annual Report Financial Data Extractor
=======================================
Extracts Balance Sheet and P&L data from annual report text
and populates an Excel file.

Usage:
    python extractor.py <pdf_file> [template.xlsx] [output.xlsx]

Requirements:
    pip install pdfplumber openpyxl
"""

import re
import sys
import logging
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# Data Model
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class BalanceSheet:
    share_capital: Optional[float] = None
    retained_earnings: Optional[float] = None
    general_reserves: Optional[float] = None
    other_equity: Optional[float] = None
    total_networth: Optional[float] = None
    accounts_payable: Optional[float] = None
    provisions_cl: Optional[float] = None
    short_term_borrowings: Optional[float] = None
    other_current_liabilities: Optional[float] = None
    other_financial_liabilities_cl: Optional[float] = None
    total_current_liabilities: Optional[float] = None
    long_term_borrowings: Optional[float] = None
    provision_ncl: Optional[float] = None
    others_ncl: Optional[float] = None
    other_financial_liabilities_ncl: Optional[float] = None
    total_non_current_liabilities: Optional[float] = None
    total_liabilities: Optional[float] = None
    bank_balance: Optional[float] = None
    cash_equivalents: Optional[float] = None
    inventory: Optional[float] = None
    investments_ca: Optional[float] = None
    loans_ca: Optional[float] = None
    accounts_receivable: Optional[float] = None
    other_current_assets: Optional[float] = None
    other_financial_assets_ca: Optional[float] = None
    total_current_assets: Optional[float] = None
    fixed_assets: Optional[float] = None
    investments_nca: Optional[float] = None
    loans_nca: Optional[float] = None
    cwip: Optional[float] = None
    other_non_current_assets: Optional[float] = None
    other_financial_assets_nca: Optional[float] = None
    deferred_tax_assets: Optional[float] = None
    total_non_current_assets: Optional[float] = None
    total_assets: Optional[float] = None

@dataclass
class ProfitAndLoss:
    revenue: Optional[float] = None
    cost_of_goods_sold: Optional[float] = None
    gross_profit: Optional[float] = None
    employee_benefits: Optional[float] = None
    interest: Optional[float] = None
    depreciation: Optional[float] = None
    other_expenses_net: Optional[float] = None
    taxes: Optional[float] = None
    net_profit: Optional[float] = None

@dataclass
class FinancialReport:
    company: str = ""
    fiscal_year: str = ""
    standalone_bs_current: BalanceSheet = field(default_factory=BalanceSheet)
    standalone_bs_prior: BalanceSheet = field(default_factory=BalanceSheet)
    standalone_pl_current: ProfitAndLoss = field(default_factory=ProfitAndLoss)
    standalone_pl_prior: ProfitAndLoss = field(default_factory=ProfitAndLoss)

# ─────────────────────────────────────────────────────────────────────────────
# Text Reader
# ─────────────────────────────────────────────────────────────────────────────

def read_source_text(path: Path) -> str:
    if pdfplumber:
        try:
            with pdfplumber.open(path) as pdf:
                pages = [p.extract_text() or "" for p in pdf.pages]
                text = "\n".join(pages)
            if text.strip():
                log.info("Extracted text via pdfplumber (%d chars)", len(text))
                return text
        except Exception:
            pass
    text = path.read_bytes().decode("utf-8", errors="replace")
    log.info("Read as plain-text file (%d chars)", len(text))
    return text

# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def parsenum(s: str) -> Optional[float]:
    s = str(s).strip()
    if s in ("-", "", "—"):
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None

def linevalues(line: str) -> list:
    nums = re.findall(r"[\d,]+\.\d+", line)
    return [float(n.replace(",", "")) for n in nums]

def parse_balance_sheet_block(block: str):
    cur = BalanceSheet()
    pri = BalanceSheet()
    lines = block.split("\n")

    def find_line(keyword, start=0):
        for ln in lines[start:]:
            if keyword.lower() in ln.lower():
                return ln.rstrip()
        return None

    def two_vals(keyword, start=0):
        ln = find_line(keyword, start)
        if not ln:
            return None, None
        vals = linevalues(ln)
        if len(vals) >= 2:
            return vals[-2], vals[-1]
        if len(vals) == 1:
            return vals[0], None
        return None, None

    cur.share_capital, pri.share_capital = two_vals("Equity Share Capital")

    ln = find_line("Other Equity")
    if ln:
        vals = linevalues(ln)
        if len(vals) >= 4:
            cur.other_equity = vals[0]; cur.total_networth = vals[1]
            pri.other_equity = vals[2]; pri.total_networth = vals[3]
        elif len(vals) >= 2:
            cur.other_equity, pri.other_equity = vals[0], vals[1]

    cl_start = next((i for i, l in enumerate(lines) if "Current Liabilities" in l and "Non" not in l), 0)
    cur.short_term_borrowings, pri.short_term_borrowings = two_vals("(i) Borrowings", cl_start)
    cur.other_current_liabilities, pri.other_current_liabilities = two_vals("Other Current Liabilities", cl_start)
    cur.other_financial_liabilities_cl, pri.other_financial_liabilities_cl = two_vals("(iii) Other Financial Liabilities", cl_start)

    msme_vals = next((linevalues(l) for l in lines[cl_start:] if "Micro Enterprises" in l), None)
    others_vals = next((linevalues(l) for l in lines[cl_start:] if "creditors other" in l.lower()), None)
    if msme_vals and others_vals:
        cur.accounts_payable = msme_vals[0] + others_vals[0]
        pri.accounts_payable = (msme_vals[1] if len(msme_vals) > 1 else 0) + (others_vals[1] if len(others_vals) > 1 else 0)

    ln = find_line("(d) Provisions", cl_start)
    if ln:
        vals = linevalues(ln)
        if len(vals) >= 4:
            cur.provisions_cl = vals[0]; cur.total_current_liabilities = vals[1]
            pri.provisions_cl = vals[2]; pri.total_current_liabilities = vals[3]

    ncl_start = next((i for i, l in enumerate(lines) if "Non Current Liabilities" in l), 0)
    cur.long_term_borrowings, pri.long_term_borrowings = two_vals("(i) Borrowing", ncl_start)
    cur.other_financial_liabilities_ncl, pri.other_financial_liabilities_ncl = two_vals("(ii) Other Financial Liabilities", ncl_start)
    cur.others_ncl, pri.others_ncl = two_vals("Other Non Current Liabilities", ncl_start)

    ln = find_line("(d) Provisions", ncl_start)
    if ln:
        m = re.search(r"\(d\)\s*Provisions\s+\d+\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)\s+(-|[\d,]+\.?\d*)\s+([\d,]+\.?\d*)", ln)
        if m:
            cur.provision_ncl = parsenum(m.group(1)) or 0.0
            cur.total_non_current_liabilities = parsenum(m.group(2))
            pri.provision_ncl = parsenum(m.group(3)) or 0.0
            pri.total_non_current_liabilities = parsenum(m.group(4))

    if cur.total_current_liabilities and cur.total_non_current_liabilities:
        cur.total_liabilities = cur.total_current_liabilities + cur.total_non_current_liabilities
    if pri.total_current_liabilities and pri.total_non_current_liabilities:
        pri.total_liabilities = pri.total_current_liabilities + pri.total_non_current_liabilities

    ca_start = next((i for i, l in enumerate(lines) if "Current Assets" in l and "Non" not in l and "Total" not in l), 0)
    cur.inventory, pri.inventory = two_vals("Inventories", ca_start)
    cur.accounts_receivable, pri.accounts_receivable = two_vals("Trade Receivable", ca_start)
    cur.cash_equivalents, pri.cash_equivalents = two_vals("Cash and cash Equivalents", ca_start)
    cur.bank_balance, pri.bank_balance = two_vals("Other Balances with Banks", ca_start)
    cur.loans_ca, pri.loans_ca = two_vals("(iv) Loans", ca_start)
    cur.other_financial_assets_ca, pri.other_financial_assets_ca = two_vals("(v) Other Financial Assets", ca_start)

    ln = find_line("Other current Assets", ca_start)
    if ln:
        vals = linevalues(ln)
        if len(vals) >= 4:
            cur.other_current_assets = vals[0]; cur.total_current_assets = vals[1]
            pri.other_current_assets = vals[2]; pri.total_current_assets = vals[3]

    nca_start = next((i for i, l in enumerate(lines) if "Non Current Assets" in l), 0)
    cur.fixed_assets, pri.fixed_assets = two_vals("Property, Plant and Equipment", nca_start)
    cur.investments_nca, pri.investments_nca = two_vals("Investment in subsidiaries", nca_start)
    cur.other_financial_assets_nca, pri.other_financial_assets_nca = two_vals("(ii) Other Financial Assets", nca_start)

    ln = find_line("Capital work in Progress", nca_start)
    if ln:
        vals = linevalues(ln)
        cur.cwip = vals[0] if vals else None
        pri.cwip = vals[1] if len(vals) > 1 else 0.0

    ln = find_line("Deferred Tax Assets (Net)", nca_start)
    if ln:
        m = re.search(r"Deferred Tax Assets.*?\s+\d+\s+(-|[\d,]+\.?\d*)\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)", ln, re.IGNORECASE)
        if m:
            cur.deferred_tax_assets = parsenum(m.group(1)) or 0.0
            cur.total_non_current_assets = parsenum(m.group(2))
            pri.deferred_tax_assets = parsenum(m.group(3))
            pri.total_non_current_assets = parsenum(m.group(4))

    cur.total_assets, pri.total_assets = two_vals("TOTAL ASSETS")
    return cur, pri


def parse_balance_sheets(text: str):
    standalone_start = text.find("BALANCE SHEET AS AT 31ST MARCH")
    consolidated_start = text.find("CONSOLIDATED BALANCE SHEET AS AT 31ST MARCH")
    sa_block = text[standalone_start:consolidated_start] if consolidated_start > standalone_start else text[standalone_start:standalone_start + 8000]
    return parse_balance_sheet_block(sa_block)


def parse_pl(text: str):
    pl_start = text.find("STATEMENT OF PROFIT AND LOSS")
    consol_pl_start = text.find("CONSOLIDATED", pl_start + 100)
    block = text[pl_start: consol_pl_start if consol_pl_start > pl_start else pl_start + 6000]

    cur = ProfitAndLoss()
    pri = ProfitAndLoss()

    def get(pattern):
        m = re.search(pattern, block, re.IGNORECASE)
        if not m:
            return None, None
        try:
            return float(m.group(1).replace(",", "")), float(m.group(2).replace(",", ""))
        except Exception:
            return None, None

    NUM = r"[\s\-]*([\d,]+\.?\d*)"
    OPT_NOTE = r"(?:\s+\d+)?"
    TWO = NUM + r"\s+" + NUM

    cur.revenue, pri.revenue = get(r"Revenue from Operation\s+" + TWO)
    mat_cur, mat_pri = get(r"Cost of Materials Consumed" + OPT_NOTE + TWO)
    con_cur, con_pri = get(r"Construction Expenses" + OPT_NOTE + TWO)
    wip_cur, wip_pri = get(r"Changes in Construction Work in Progress" + OPT_NOTE + r"\s*\(?(\d[\d,]*\.?\d*)\)?\s+\(?(\d[\d,]*\.?\d*)\)?")

    if mat_cur is not None:
        cur.cost_of_goods_sold = (mat_cur or 0) + (con_cur or 0) - (wip_cur or 0)
        pri.cost_of_goods_sold = (mat_pri or 0) + (con_pri or 0) - (wip_pri or 0)

    if cur.revenue and cur.cost_of_goods_sold:
        cur.gross_profit = cur.revenue - cur.cost_of_goods_sold
    if pri.revenue and pri.cost_of_goods_sold:
        pri.gross_profit = pri.revenue - pri.cost_of_goods_sold

    cur.employee_benefits, pri.employee_benefits = get(r"Employee Benefits Expenses" + OPT_NOTE + TWO)
    cur.interest, pri.interest = get(r"Finance Costs" + OPT_NOTE + TWO)
    cur.depreciation, pri.depreciation = get(r"Depreciation and Amortisation(?:\s+Expense)?" + OPT_NOTE + TWO)
    cur.other_expenses_net, pri.other_expenses_net = get(r"Other Expenses" + OPT_NOTE + TWO)

    m_tax = re.search(r"Deferred tax\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)\s+\(?([\d,]+\.?\d*)\)?\s+([\d,]+\.?\d*)", block)
    if m_tax:
        cur.taxes = float(m_tax.group(2).replace(",", ""))
        pri.taxes = float(m_tax.group(4).replace(",", ""))

    cur.net_profit, pri.net_profit = get(r"Profit for the year.*?" + TWO)
    return cur, pri


# ─────────────────────────────────────────────────────────────────────────────
# Excel Writer
# ─────────────────────────────────────────────────────────────────────────────

BLUE = "FF0070C0"
GREEN = "FF00B050"
SECTION_BG = "FFBDD7EE"


def populate_excel(output_path: Path, report: FinancialReport):
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Summary"

    bs_c = report.standalone_bs_current
    bs_p = report.standalone_bs_prior
    pl_c = report.standalone_pl_current
    pl_p = report.standalone_pl_prior

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    def hdr(row, label):
        c = ws.cell(row, 1, label)
        c.font = Font(name="Arial", bold=True, size=10, color="FF1F4E79")
        c.fill = PatternFill("solid", fgColor=SECTION_BG)
        ws.cell(row, 2).fill = PatternFill("solid", fgColor=SECTION_BG)
        ws.cell(row, 3).fill = PatternFill("solid", fgColor=SECTION_BG)

    def val(row, label, cy, py, total=False):
        lc = ws.cell(row, 1, label)
        cc = ws.cell(row, 2, round(cy, 2) if cy is not None else "-")
        pc = ws.cell(row, 3, round(py, 2) if py is not None else "-")
        lc.font = Font(name="Arial", bold=total, size=10)
        for c in [cc, pc]:
            c.alignment = Alignment(horizontal="right")
            c.number_format = '#,##0.00'
            c.font = Font(name="Arial", bold=total, size=10, color=GREEN if total else BLUE)

    # Header
    ws.cell(1, 1, report.company).font = Font(name="Arial", bold=True, size=12)
    ws.cell(2, 1, "Financial Summary (INR in Lakhs)").font = Font(name="Arial", bold=True, size=10)
    ws.cell(2, 2, "FY Current").font = Font(name="Arial", bold=True, size=10, color=BLUE)
    ws.cell(2, 3, "FY Prior").font = Font(name="Arial", bold=True, size=10, color=BLUE)

    rows = [
        (3,  "BALANCE SHEET", None, None, False, True),
        (4,  "NETWORTH", None, None, False, True),
        (5,  "Share Capital", bs_c.share_capital, bs_p.share_capital, False, False),
        (6,  "Other Equity", bs_c.other_equity, bs_p.other_equity, False, False),
        (7,  "Total Networth", bs_c.total_networth, bs_p.total_networth, True, False),
        (8,  "CURRENT LIABILITIES", None, None, False, True),
        (9,  "Short Term Borrowings", bs_c.short_term_borrowings, bs_p.short_term_borrowings, False, False),
        (10, "Accounts Payable", bs_c.accounts_payable, bs_p.accounts_payable, False, False),
        (11, "Other Current Liabilities", bs_c.other_current_liabilities, bs_p.other_current_liabilities, False, False),
        (12, "Other Financial Liabilities", bs_c.other_financial_liabilities_cl, bs_p.other_financial_liabilities_cl, False, False),
        (13, "Provisions", bs_c.provisions_cl, bs_p.provisions_cl, False, False),
        (14, "Total Current Liabilities", bs_c.total_current_liabilities, bs_p.total_current_liabilities, True, False),
        (15, "NON-CURRENT LIABILITIES", None, None, False, True),
        (16, "Long Term Borrowings", bs_c.long_term_borrowings, bs_p.long_term_borrowings, False, False),
        (17, "Other Financial Liabilities", bs_c.other_financial_liabilities_ncl, bs_p.other_financial_liabilities_ncl, False, False),
        (18, "Others", bs_c.others_ncl, bs_p.others_ncl, False, False),
        (19, "Provisions", bs_c.provision_ncl, bs_p.provision_ncl, False, False),
        (20, "Total Non-Current Liabilities", bs_c.total_non_current_liabilities, bs_p.total_non_current_liabilities, True, False),
        (21, "TOTAL LIABILITIES", bs_c.total_liabilities, bs_p.total_liabilities, True, False),
        (22, "CURRENT ASSETS", None, None, False, True),
        (23, "Cash & Cash Equivalents", bs_c.cash_equivalents, bs_p.cash_equivalents, False, False),
        (24, "Bank Balance", bs_c.bank_balance, bs_p.bank_balance, False, False),
        (25, "Inventory", bs_c.inventory, bs_p.inventory, False, False),
        (26, "Accounts Receivable", bs_c.accounts_receivable, bs_p.accounts_receivable, False, False),
        (27, "Loans", bs_c.loans_ca, bs_p.loans_ca, False, False),
        (28, "Other Financial Assets", bs_c.other_financial_assets_ca, bs_p.other_financial_assets_ca, False, False),
        (29, "Other Current Assets", bs_c.other_current_assets, bs_p.other_current_assets, False, False),
        (30, "Total Current Assets", bs_c.total_current_assets, bs_p.total_current_assets, True, False),
        (31, "NON-CURRENT ASSETS", None, None, False, True),
        (32, "Fixed Assets (PPE)", bs_c.fixed_assets, bs_p.fixed_assets, False, False),
        (33, "Investments", bs_c.investments_nca, bs_p.investments_nca, False, False),
        (34, "Loans", bs_c.loans_nca, bs_p.loans_nca, False, False),
        (35, "Capital Work-in-Progress", bs_c.cwip, bs_p.cwip, False, False),
        (36, "Deferred Tax Assets", bs_c.deferred_tax_assets, bs_p.deferred_tax_assets, False, False),
        (37, "Total Non-Current Assets", bs_c.total_non_current_assets, bs_p.total_non_current_assets, True, False),
        (38, "TOTAL ASSETS", bs_c.total_assets, bs_p.total_assets, True, False),
        (40, "P&L STATEMENT", None, None, False, True),
        (41, "Revenue from Operations", pl_c.revenue, pl_p.revenue, False, False),
        (42, "Cost of Goods Sold", pl_c.cost_of_goods_sold, pl_p.cost_of_goods_sold, False, False),
        (43, "Gross Profit", pl_c.gross_profit, pl_p.gross_profit, True, False),
        (44, "Employee Benefits Expense", pl_c.employee_benefits, pl_p.employee_benefits, False, False),
        (45, "Finance Costs (Interest)", pl_c.interest, pl_p.interest, False, False),
        (46, "Depreciation & Amortisation", pl_c.depreciation, pl_p.depreciation, False, False),
        (47, "Other Expenses", pl_c.other_expenses_net, pl_p.other_expenses_net, False, False),
        (48, "Total Tax", pl_c.taxes, pl_p.taxes, False, False),
        (49, "Net Profit for the Year", pl_c.net_profit, pl_p.net_profit, True, False),
    ]

    for (row, label, cy, py, total, is_hdr) in rows:
        if is_hdr:
            hdr(row, label)
        else:
            val(row, label, cy, py, total)

    wb.save(output_path)
    log.info("Saved: %s", output_path)


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def run(source_file: str, output_file: str, company: str = ""):
    source = Path(source_file)
    output = Path(output_file)

    log.info("Reading: %s", source)
    text = read_source_text(source)

    log.info("Parsing Balance Sheet...")
    bs_cur, bs_pri = parse_balance_sheets(text)

    log.info("Parsing P&L...")
    pl_cur, pl_pri = parse_pl(text)

    if not company:
        company = source.stem.replace("_", " ").replace("-", " ")

    report = FinancialReport(
        company=company,
        fiscal_year="FY Current",
        standalone_bs_current=bs_cur,
        standalone_bs_prior=bs_pri,
        standalone_pl_current=pl_cur,
        standalone_pl_prior=pl_pri,
    )

    log.info("Writing Excel...")
    populate_excel(output, report)

    print("\n" + "=" * 60)
    print("EXTRACTION SUMMARY — " + company)
    print("=" * 60)

    def row(label, cy, py):
        cy_s = f"{cy:>12,.2f}" if cy is not None else f"{'N/A':>12}"
        py_s = f"{py:>12,.2f}" if py is not None else f"{'N/A':>12}"
        print(f"{label:<38} {cy_s} {py_s}")

    print(f"\n{'Field':<38} {'Current':>12} {'Prior':>12}")
    print("-" * 64)
    row("  Total Assets", bs_cur.total_assets, bs_pri.total_assets)
    row("  Total Networth", bs_cur.total_networth, bs_pri.total_networth)
    row("  Total Liabilities", bs_cur.total_liabilities, bs_pri.total_liabilities)
    row("  Revenue", pl_cur.revenue, pl_pri.revenue)
    row("  Net Profit", pl_cur.net_profit, pl_pri.net_profit)
    print("=" * 64)
    print(f"\nOutput: {output}")


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else "annual_report.pdf"
    out = sys.argv[2] if len(sys.argv) > 2 else "financial_summary.xlsx"
    co  = sys.argv[3] if len(sys.argv) > 3 else ""
    run(src, out, co)
